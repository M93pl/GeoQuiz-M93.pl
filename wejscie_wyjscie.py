import openpyxl
from openpyxl import Workbook, load_workbook



def wprowadzenie_danych1(dane_wejściowe):
	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active
	dane_wej=dane_wejściowe.capitalize()
	arkusz['B2'].value=(dane_wej)
	baza_danych.save('baza.xlsx')

def wyjscie_danych1():
	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active
	return(arkusz['B2'].value)
	baza_danych.save('baza.xlsx')

def usuniecie_danych1():
	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active
	arkusz['B2'].value=None
	baza_danych.save('baza.xlsx')

	

def wprowadzenie_danych2(dane_wejściowe):
	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active
	arkusz['D3'].value=(dane_wejściowe)
	baza_danych.save('baza.xlsx')

def wyjscie_danych2():
	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active
	return(arkusz['D3'].value)
	baza_danych.save('baza.xlsx')

def usuniecie_danych2():
	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active
	arkusz['D3'].value=None
	baza_danych.save('baza.xlsx')




def wprowadzenie_danych3(dane_wejściowe):
	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active
	arkusz['C3'].value=(dane_wejściowe)
	baza_danych.save('baza.xlsx')

def wyjscie_danych3():
	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active
	return(arkusz['C3'].value)
	baza_danych.save('baza.xlsx')

def usuniecie_danych3():
	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active
	arkusz['C3'].value=None
	baza_danych.save('baza.xlsx')



def wprowadzenie_danych4(dane_wejściowe):
	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active
	arkusz['D4'].value=(dane_wejściowe)
	baza_danych.save('baza.xlsx')

def wyjscie_danych4():
	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active
	return(arkusz['D4'].value)
	baza_danych.save('baza.xlsx')

def usuniecie_danych4():
	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active
	arkusz['D4'].value=None
	baza_danych.save('baza.xlsx')



def wprowadzenie_danych5(dane_wejściowe):
	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active
	arkusz['C4'].value=(dane_wejściowe)
	baza_danych.save('baza.xlsx')

def wyjscie_danych5():
	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active
	return(arkusz['C4'].value)
	baza_danych.save('baza.xlsx')

def usuniecie_danych5():
	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active
	arkusz['C4'].value=None
	baza_danych.save('baza.xlsx')


