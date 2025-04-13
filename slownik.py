from wejscie_wyjscie import *
from os import system
from random import *
from wyniki import *
from tkinter import *
import openpyxl
from openpyxl import Workbook, load_workbook
from tkinter import ttk

# TWORZENIE OKNA I PODSTAWOWE FUNKCJE

def ABC(okno):
	okno.withdraw()
	okno_ABC=Toplevel()
	okno_ABC.focus_set()
	okno_ABC.focus_force()
	okno_ABC.geometry('600x500')
	okno_ABC.title('Stolice')
	okno_ABC.resizable(False, False)

	class PytanieSzablon ():
		def __init__(self, pytanie, odpowiedzi, prawidlowa):
			self.pytanie=pytanie
			self.odpowiedzi=odpowiedzi
			self.prawidlowa=prawidlowa


	pyt1=PytanieSzablon('W którym roku założono Poleski Park Narodowy?',\
	['1990','1992','2003'], '1990')
	pyt2=PytanieSzablon('Najbardziej kasowy film wszechczasów to:',\
	['Avatar','Titanic','Szczęki'], 'Avatar')
	pyt3=PytanieSzablon('Biblijny Jakub był synem:',\
	['Nachora','Izaaka','Abrahama'], 'Izaaka')
	pyt4=PytanieSzablon('W którym roku wykonano w Polsce\nkarę śmierci po raz ostatni?',\
	['1992','1972','1988'], '1988')
	pyt5=PytanieSzablon('Jaką nazwę nosi język programowania\nw jakim napisano ten program?',\
	['C++','Python','Java'], 'Python')
	pyt6=PytanieSzablon('Kto jest królem Anglii?',\
	['Wilhelm V','Elżbieta II','Karol III'], 'Karol III')
	pyt7=PytanieSzablon('Które z jezior znajduje się w Niemczech?',\
	['Chiemsee','Sommen','Achensee'], 'Chiemsee')
	pyt8=PytanieSzablon('Której gry komputerowej nie stworzyli Polacy?',\
	['Cyberpunk 2077','Minecraft','The Witcher'], 'Minecraft')
	pyt9=PytanieSzablon('Kto jest marszałkiem senatu?',\
	['W. Czarzasty','S. Hołownia','M. Kidawa-Błońska'], 'M. Kidawa-Błońska')
	pyt10=PytanieSzablon('Która z partii nie wygrała nigdy wyborów w Polsce?',\
	['PSL','SLD','PiS'], 'PSL')
	pyt11=PytanieSzablon('Kim była Łarysa Łatynina?',\
	['Pływaczką','Gimnastyczką','Sprinterką'], 'Gimnastyczką')
	pyt12=PytanieSzablon('Które spośród poniższych państw\nzdobyło najwięcej medali olimpijskich?',\
	['Hiszpania','Polska','Węgry'], 'Hiszpania')
	pyt13=PytanieSzablon('Moment to jednostka czasu określająca:',\
	['1/100 sekundy','1/8 sekundy','dwie sekundy'], '1/100 sekundy')
	pyt14=PytanieSzablon('Państwo Brazylia nosi nazwę po:',\
	['wyrazie w języku portugalskim','gatunku drzewa','nazwisku swojego odkrywcy'],'gatunku drzewa')
	pyt15=PytanieSzablon('Z jakiego języka wywopdzi się słowo "alkohol"?',\
	['tureckiego','ukraińskiego','arabskiego'], 'arabskiego')
	pyt16=PytanieSzablon('Kto był pierwszym cesarzem Rzymu?',\
	['Oktawian August','Juliusz Cezar','Marek Aureliusz'], 'Oktawian August')
	pyt17=PytanieSzablon('W jakim wieku zmarł Napoleon Bonaparte?',\
	['51 lat','48 lat','62 lata'], '51 lat')
	pyt18=PytanieSzablon('Ile wyścigów Formuły 1 wygrał Robert Kubica?',\
	['dwa','jeden','żadnego'], 'jeden')
	pyt19=PytanieSzablon('Z jakiego kraju pochodzi Cristiano Ronaldo?',\
	['z Brazylii','z Portugalii','z Włoch'], 'z Portugalii')
	pyt20=PytanieSzablon('Z jakiego kontynentu pochodzi tytoń?',\
	['z Azji','z Afryki','z Ameryki Południowej'], 'z Ameryki Południowej')
	pyt21=PytanieSzablon('W krórym wieku założono Teatr\nim. Juliusza Osterwy w Lublinie?',\
	['w XVIII wieku','w XIX wieku','w XX wieku'], 'w XX wieku')
	pyt22=PytanieSzablon('Czego jest więcej na ziemi?',\
	['bydła hodowlanego', 'dzikich zwierząt'], 'bydła hodowlanego')
	pyt23=PytanieSzablon('Co waży więcej?',\
	['15 uncji', '1 funt'], '1 funt')
	pyt24=PytanieSzablon('Z jakiego kraju pochodzi Elon Musk?',\
	['z Kanady', 'z RPA', 'z USA'], 'z RPA')
	pyt25=PytanieSzablon('W którym roku zmarł Stanisław Lem?',\
	['1985', '1994', '2006'], '2006')
	pyt26=PytanieSzablon('Czy Republika Czeska jest w Unii Europejskiej?',\
	['Tak', 'Nie'], 'Tak')
	pyt27=PytanieSzablon('Czy Nil uchodzi do Morza Czarnego?',\
	['Tak', 'Nie'], 'Nie')
	pyt28=PytanieSzablon('Czy Wisła to najdłuższa rzeka w Polsce?',\
	['Tak', 'Nie'], 'Tak')
	pyt29=PytanieSzablon('Czy Euro jest walutą obowiązującą w Szwecji?', \
	['Tak', 'Nie'], 'Nie')
	pyt30=PytanieSzablon('Czy arachnofobia polega na lęku przed wysokością?',\
	['Tak', 'Nie'], 'Nie')
	pyt31=PytanieSzablon('Góra Elbrus to najwyższy szczyt Kaukazu.',\
	['Prawda', 'Fałsz'], 'Prawda')
	pyt32=PytanieSzablon('Rozpiętość skrzydeł czapli białej nie przekracza 1 metra.',\
	['Prawda', 'Fałsz'], 'Fałsz')
	pyt33=PytanieSzablon('Holandia to królestwo.',\
	['Prawda', 'Fałsz'], 'Prawda')
	pyt34=PytanieSzablon('Tybet to niepodległe państwo.', \
	['Prawda', 'Fałsz'], 'Fałsz')
	pyt35=PytanieSzablon('Erytrocyty transportują tlen.', \
	['Prawda', 'Fałsz'], 'Prawda')
	pyt36=PytanieSzablon('Słowo "rupia" oznacza:',\
	['złoto', 'srebro'], 'srebro')
	pyt37=PytanieSzablon('Miasto Meksyk ma około 3 milionów mieszkańców.',\
	['Prawda', 'Fałsz'], 'Fałsz')
	pyt38=PytanieSzablon('W którym roku odbyła się bitwa pod Grunwaldem?', \
	['1405', '1410', '1450'], '1410')
	pyt39=PytanieSzablon('Który z podanych krajów nie graniczy z Polską?', \
	['Litwa', 'Ukraina', 'Austria'], 'Austria')
	pyt40=PytanieSzablon('Która z planet Układu Słonecznego jest największa?', \
	['Saturn', 'Mars', 'Jowisz'], 'Jowisz')
	pyt41=PytanieSzablon('Która z postaci literackich\nstworzona została przez J.R.R. Tolkiena?',\
	['Harry Potter', 'Frodo Baggins', 'Sherlock Holmes'], 'Frodo Baggins')
	pyt42=PytanieSzablon('Kto jest autorem sztuki "Romeo i Julia"?', \
	['William Shakespeare', 'Oscar Wilde', 'Jane Austen'], 'William Shakespeare')
	pyt43=PytanieSzablon('W którym roku miała miejsce\npierwsza wyprawa człowieka na Księżyc?',\
	['1959', '1969', '1979'], '1969')
	pyt44=PytanieSzablon('Który z oceanów jest największy\npod względem powierzchni?', \
	['Ocean Atlantycki', 'Ocean Spokojny', 'Ocean Indyjski'], 'Ocean Spokojny')
	pyt45=PytanieSzablon('Która z rzek przepływa\nprzez najwięcej krajów?', \
	['Nil', 'Amazonka', 'Dunaj'], 'Nil')
	pyt46=PytanieSzablon('Która z poniższych górskich\nformacji znajduje się w Afryce?',\
	['Alpy', 'Himalaje', 'Góry Atlas'], 'Góry Atlas')
	pyt47=PytanieSzablon('Który z podanych krajów leży na Półwyspie Iberyjskim?',\
	['Włochy', 'Portugalia', 'Grecja'], 'Portugalia')
	pyt48=PytanieSzablon('Który z podanych miast jest stolicą Danii?', \
	['Sztokholm', 'Oslo', 'Kopenhaga'], 'Kopenhaga')
	pyt49=PytanieSzablon('Który z podanych języków programowania jest najstarszy?', \
	['C++', 'Python', 'Java'], 'C++')
	pyt50=PytanieSzablon('Który z podanych zespołów muzycznych\nNIE pochodzi z Anglii?', \
	['U2', 'The Beatles', 'Led Zeppelin'], 'U2')
	pyt51=PytanieSzablon('Który z podanych polskich pisarzy napisał\n"Ferdydurke"?', \
	['Juliusz Słowacki', 'Henryk Sienkiewicz', 'Witold Gombrowicz'], 'Witold Gombrowicz')
	pyt52=PytanieSzablon('Który z podanych aktorów zagrał postać Jamesa Bonda?', \
	['Sean Connery', 'Arnold Schwarzenegger', 'Harrison Ford'], 'Sean Connery')
	pyt53=PytanieSzablon('Które z podanych zwierząt jest\nnajwiększym lądowym ssakiem?', \
	['Słoń', 'Nosorożec', 'Hipopotam'], 'Słoń')
	pyt54=PytanieSzablon('Która z planet jest najbliżej Słońca?',\
	['Wenus', 'Mars', 'Saturn'], 'Wenus')
	pyt55=PytanieSzablon('Które z podanych zwierząt to symbol\nkrólewski w Wielkiej Brytanii?', \
	['Lew', 'Orzeł', 'Jednorożec'], 'Lew')
	pyt56=PytanieSzablon('Który z podanych państw jest największym\nproducentem kawy na świecie?',\
	['Brazylia', 'Kolumbia', 'Wietnam'], 'Brazylia')
	pyt57=PytanieSzablon('Który z podanych zespołów muzycznych\npochodzi z Irlandii?', \
	['U2', 'ABBA', 'Queen'], 'U2')
	pyt58=PytanieSzablon('Który z podanych wynalazców wynalazł żarówkę?',\
	['Nikola Tesla', 'Albert Einstein', 'Thomas Edison'], 'Thomas Edison')


	lista_pytan=[pyt1, pyt2, pyt3, pyt4, pyt5, pyt6, pyt7, pyt8, pyt9, pyt10, pyt11, pyt12,\
	pyt13, pyt14, pyt15, pyt16, pyt17, pyt18, pyt19, pyt20, pyt21, pyt22, pyt23, pyt24,\
	pyt25, pyt26, pyt27, pyt28, pyt29, pyt30, pyt31, pyt32, pyt33, pyt34, pyt35,\
	pyt36, pyt37, pyt38, pyt39, pyt40, pyt41, pyt42, pyt43, pyt44, pyt45, pyt46,\
	pyt47, pyt48, pyt49, pyt50, pyt51, pyt52, pyt53, pyt54, pyt55, pyt56, pyt57, pyt58]




#	 GRA GŁÓWNA


	def start(x):

		global count
		count+=1
		if count<16:
			wart1=StringVar()

			style=ttk.Style()
			style.configure("S1.Toolbutton",padding=(15,15), font=('Arial',14))
			style.configure("S2.Toolbutton",padding=(15,15), font=('Arial',14),foreground='green')
			x.destroy()
			rama1=Frame(okno_ABC)
			rama1.pack()

			pytanieLos=choice(lista_pytan)

			pytajnik=Label(rama1,text=pytanieLos.pytanie,font=("Arial",16,'bold'))
			pytajnik.pack(pady=25)
			shuffle(pytanieLos.odpowiedzi)

			dobrze=choice(['Prawidłowo!','Zgadza się!','Tak!','Brawo!','Dobrze!',\
					'Świetnie!', 'Doskonale!', 'Rewelacyjnie!', 'Fantastycznie!', 'Super!',\
					'Bezapelacyjnie!', 'Oczywiście!', 'Jasne!', 'Absolutnie tak!',\
					'Bez wątpienia!', 'Okej!', 'Nieźle!'])
			zle=choice(['Źle!','Odpowiedź nieprawidłowa!','Nie!',\
					'Niestety nie.', 'Błąd :/', 'Pudło...', 'Niestety, to nie to :(', 'Nie tym razem.'])

			def activate():
				sprawdz.config(state='active')

			if len(pytanieLos.odpowiedzi)==3:
				opcjaA=ttk.Radiobutton(rama1,style='S1.Toolbutton',command=activate,\
					text=pytanieLos.odpowiedzi[0], variable=wart1,value=pytanieLos.odpowiedzi[0])
				opcjaA.pack()
				opcjaB=ttk.Radiobutton(rama1,style='S1.Toolbutton',command=activate,\
					text=pytanieLos.odpowiedzi[1], variable=wart1,value=pytanieLos.odpowiedzi[1])
				opcjaB.pack()
				opcjaC=ttk.Radiobutton(rama1,style='S1.Toolbutton',command=activate,\
					text=pytanieLos.odpowiedzi[2], variable=wart1,value=pytanieLos.odpowiedzi[2])
				opcjaC.pack()
				




			if len(pytanieLos.odpowiedzi)==2:
				opcjaA=ttk.Radiobutton(rama1,style='S1.Toolbutton',command=activate,\
					text=pytanieLos.odpowiedzi[0], variable=wart1,value=pytanieLos.odpowiedzi[0])
				opcjaA.pack()
				opcjaB=ttk.Radiobutton(rama1,style='S1.Toolbutton',command=activate,\
					text=pytanieLos.odpowiedzi[1], variable=wart1,value=pytanieLos.odpowiedzi[1])
				opcjaB.pack()



			def check():
				global pkt
				if pytanieLos.odpowiedzi[0]==pytanieLos.prawidlowa:
					opcjaA.config(style='S2.Toolbutton')
				elif pytanieLos.odpowiedzi[1]==pytanieLos.prawidlowa:
					opcjaB.config(style='S2.Toolbutton')
				elif pytanieLos.odpowiedzi[2]==pytanieLos.prawidlowa:
					opcjaC.config(style='S2.Toolbutton')

				if wart1.get()==pytanieLos.prawidlowa:
					pkt+=1
					pytajnik.config(text=f'{dobrze}\n\nMasz w sumie:\n{pkt}pkt.',\
						fg='green',font=("Arial",15,'bold'))
					sprawdz.config(text='Następne pytanie',state='active', \
						command=lambda:start(rama1))
					lista_pytan.remove(pytanieLos)
					
				else:
					pytajnik.config(text=f'{zle}\n\nMasz w sumie:\n{pkt}pkt.',\
						fg='red',font=("Arial",15,'bold'))
					sprawdz.config(text='Następne pytanie',state='active', \
						command=lambda:start(rama1))
					lista_pytan.remove(pytanieLos)
					



			sprawdz=Button(rama1,state='disabled',font=("Arial",16),text='Sprawdź!',\
				command=check,borderwidth=5)
			sprawdz.pack(pady=18,ipadx=4)


#	ZAKOŃCZENIE:

		else:
			x.destroy()
			ramka_gra=Frame(okno_ABC)
			ramka_gra.pack()
			zakonczenie=Label(ramka_gra,text=f'Koniec!\
\nTwój wynik to: {pkt} pkt.', font=('Arial', 20)).pack(pady=15)

			if wyjscie_danych5()==None or wyjscie_danych4()==None or int(wyjscie_danych4())<int(pkt):
				zakonczenie2=Label(ramka_gra,font=('Arial',15),text=f'Rewelacja {wyjscie_danych1()}!\
\nTo najlepszy dotychczasowy wynik!').pack(pady=15)
				wprowadzenie_danych4(str(pkt))
				wprowadzenie_danych5(str(wyjscie_danych1()))
				ustal_rank_abc(pkt)

			elif int(wyjscie_danych4())==int(pkt):
				zakonczenie2=Label(ramka_gra,font=('Arial',15),text=f'Brawo {wyjscie_danych1()}\
\n{pkt}pkt. to najwyższy wynik!\nZdobywasz tyle samo punktów co {wyjscie_danych5()}!!').pack(pady=15)
				ustal_rank_abc(pkt)

			elif int(wyjscie_danych4())>int(pkt):
				zakonczenie2=Label(ramka_gra,font=('Arial',15),text=f'Najlepszy wynik dotychczas to:\
\n{wyjscie_danych4()}pkt.\nOsiągnięte przez użytkownika:\n{wyjscie_danych5()}').pack(pady=15)
				ustal_rank_abc(pkt)
			

			def best():
				baza_danych=load_workbook('baza.xlsx')
				arkusz=baza_danych.active
				wyniki.config(state='disabled')
				g1=arkusz['B19'].value
				g2=arkusz['B20'].value
				g3=arkusz['B21'].value
				g4=arkusz['B22'].value
				g5=arkusz['B23'].value

				p1=arkusz['A19'].value
				p2=arkusz['A20'].value
				p3=arkusz['A21'].value
				p4=arkusz['A22'].value
				p5=arkusz['A23'].value


				okno_info=Tk()
				okno_info.geometry('300x200')
				okno_info.title('Highscores')
				baza_danych=load_workbook('baza.xlsx')
				info_napis=Label(okno_info,padx=20, text=f'Najlepsze dotychczasowe wyniki:\
\n\
\n1. {g1}   {p1}pkt.\
\n2. {g2}   {p2}pkt.\
\n3. {g3}   {p3}pkt.\
\n4. {g4}   {p4}pkt.\
\n5. {g5}   {p5}pkt.', font=('Arial', 12))
				info_napis.pack()
				baza_danych.close()


				def powrot_infoF():
					wyniki.config(state='active')
					okno_info.quit()
					okno_info.destroy()

				powrot_info=Button(okno_info,text='OK',command=powrot_infoF, font=('Arial', 18))
				powrot_info.pack(pady=15, ipadx=25)
				okno_info.mainloop()

			wyniki=Button(okno_ABC,text='Wyświetl najlepsze wyniki', command=best, font=('Arial', 18))
			wyniki.pack(pady=10, side='bottom')



	global count
	count=0
	global pkt
	pkt=0

	rama1=Frame(okno_ABC)
	rama1.pack()


	powitanie=Label(rama1,text=f'{wyjscie_danych1()} postaraj się odpowiedzieć\
\nna jak największą liczbę pytań!',font=("Arial",20)).pack(pady=45)

	guzik1=Button(rama1,text='Zaczynamy!', command=lambda:start(rama1), \
		font=("Arial",24)).pack(pady=40)




	def cls_(nazwa_okna):
		for widgety in nazwa_okna.winfo_children():
			widgety.pack_forget()

	def cls_bez_zamknij(nazwa_okna, bez_zamknij):
		for widgety in nazwa_okna.winfo_children():
			if widgety is not bez_zamknij:
				widgety.pack_forget()



# PĘTLA OKNA I WYJŚCIE Z OKNA

	def powrot():
		okno_ABC.destroy()
		okno.deiconify()

	zamknij=Button(okno_ABC,text='Powrót do menu', borderwidth=4, \
		command=powrot, font=("Arial",18))
	zamknij.pack(padx=15,pady=15,side='bottom')
	okno_ABC.mainloop()









