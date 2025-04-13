import os
import openpyxl
from wejscie_wyjscie import *
from openpyxl import Workbook, load_workbook


os.system('cls')




def ustal_rank(punkty):

	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active

	lista_wyników=[
	[arkusz['A11'].value,'pkt.\t',arkusz['B11'].value],
	[arkusz['A12'].value,'pkt.\t',arkusz['B12'].value],
	[arkusz['A13'].value,'pkt.\t',arkusz['B13'].value],
	[arkusz['A14'].value,'pkt.\t',arkusz['B14'].value],
	[arkusz['A15'].value,'pkt.\t',arkusz['B15'].value]
	]

	def sort1(sor1):
		return sor1[0] 

	lista_wyników.append([punkty,'pkt.\t',wyjscie_danych1()])
	lista_wyników.sort(key=sort1,reverse=True)
	miejsca=['1.','2.','3.','4.','5.']
	lista_wyników.remove(min(lista_wyników))


	arkusz['A11'].value=lista_wyników[0][0]
	arkusz['A12'].value=lista_wyników[1][0]
	arkusz['A13'].value=lista_wyników[2][0]
	arkusz['A14'].value=lista_wyników[3][0]
	arkusz['A15'].value=lista_wyników[4][0]

	arkusz['B11'].value=lista_wyników[0][2]
	arkusz['B12'].value=lista_wyników[1][2]
	arkusz['B13'].value=lista_wyników[2][2]
	arkusz['B14'].value=lista_wyników[3][2]
	arkusz['B15'].value=lista_wyników[4][2]

	baza_danych.save('baza.xlsx')



def usun_rank():

	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active

	arkusz['A11'].value=0
	arkusz['A12'].value=0
	arkusz['A13'].value=0
	arkusz['A14'].value=0
	arkusz['A15'].value=0

	arkusz['B11'].value='<zbyt mało prób>'
	arkusz['B12'].value='<zbyt mało prób>'
	arkusz['B13'].value='<zbyt mało prób>'
	arkusz['B14'].value='<zbyt mało prób>'
	arkusz['B15'].value='<zbyt mało prób>'

	baza_danych.save('baza.xlsx')


def ustal_rank_abc(punk):

	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active

	lista_wyników=[
	[arkusz['A19'].value,'pkt.\t',arkusz['B19'].value],
	[arkusz['A20'].value,'pkt.\t',arkusz['B20'].value],
	[arkusz['A21'].value,'pkt.\t',arkusz['B21'].value],
	[arkusz['A22'].value,'pkt.\t',arkusz['B22'].value],
	[arkusz['A23'].value,'pkt.\t',arkusz['B23'].value]
	]

	def sort1(sor1):
		return sor1[0] 

	lista_wyników.append([punk,'pkt.\t',wyjscie_danych1()])
	lista_wyników.sort(key=sort1,reverse=True)
	miejsca=['1.','2.','3.','4.','5.']
	lista_wyników.remove(min(lista_wyników))

	arkusz['A19'].value=lista_wyników[0][0]
	arkusz['A20'].value=lista_wyników[1][0]
	arkusz['A21'].value=lista_wyników[2][0]
	arkusz['A22'].value=lista_wyników[3][0]
	arkusz['A23'].value=lista_wyników[4][0]

	arkusz['B19'].value=lista_wyników[0][2]
	arkusz['B20'].value=lista_wyników[1][2]
	arkusz['B21'].value=lista_wyników[2][2]
	arkusz['B22'].value=lista_wyników[3][2]
	arkusz['B23'].value=lista_wyników[4][2]

	baza_danych.save('baza.xlsx')



def usun_rank_abc():

	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active

	arkusz['A19'].value=0
	arkusz['A20'].value=0
	arkusz['A21'].value=0
	arkusz['A22'].value=0
	arkusz['A23'].value=0

	arkusz['B19'].value='<zbyt mało prób>'
	arkusz['B20'].value='<zbyt mało prób>'
	arkusz['B21'].value='<zbyt mało prób>'
	arkusz['B22'].value='<zbyt mało prób>'
	arkusz['B23'].value='<zbyt mało prób>'

	baza_danych.save('baza.xlsx')



def ustal_rank_atlas(punkty):

	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active

	lista_wyników=[
	[arkusz['A29'].value,'%\t',arkusz['B29'].value],
	[arkusz['A30'].value,'%\t',arkusz['B30'].value],
	[arkusz['A31'].value,'%\t',arkusz['B31'].value],
	[arkusz['A32'].value,'%\t',arkusz['B32'].value],
	[arkusz['A33'].value,'%\t',arkusz['B33'].value]
	]

	def sort1(sor1):
		return sor1[0] 

	lista_wyników.append([punkty,'%\t',wyjscie_danych1()])
	lista_wyników.sort(key=sort1,reverse=True)
	miejsca=['1.','2.','3.','4.','5.']
	lista_wyników.remove(min(lista_wyników))


	arkusz['A29'].value=lista_wyników[0][0]
	arkusz['A30'].value=lista_wyników[1][0]
	arkusz['A31'].value=lista_wyników[2][0]
	arkusz['A32'].value=lista_wyników[3][0]
	arkusz['A33'].value=lista_wyników[4][0]

	arkusz['B29'].value=lista_wyników[0][2]
	arkusz['B30'].value=lista_wyników[1][2]
	arkusz['B31'].value=lista_wyników[2][2]
	arkusz['B32'].value=lista_wyników[3][2]
	arkusz['B33'].value=lista_wyników[4][2]

	baza_danych.save('baza.xlsx')



def usun_rank_atlas():

	baza_danych=load_workbook('baza.xlsx')
	arkusz=baza_danych.active

	arkusz['A29'].value=0
	arkusz['A30'].value=0
	arkusz['A31'].value=0
	arkusz['A32'].value=0
	arkusz['A33'].value=0

	arkusz['B29'].value='<zbyt mało prób>'
	arkusz['B30'].value='<zbyt mało prób>'
	arkusz['B31'].value='<zbyt mało prób>'
	arkusz['B32'].value='<zbyt mało prób>'
	arkusz['B33'].value='<zbyt mało prób>'

	baza_danych.save('baza.xlsx')









