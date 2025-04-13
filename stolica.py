from wejscie_wyjscie import *
from wyniki import *
from os import system
from random import *
from tkinter import *
import openpyxl
from openpyxl import Workbook, load_workbook

# TWORZENIE OKNA I PODSTAWOWE FUNKCJE

def stolice(okno):

	okno.withdraw()
	okno_stol=Toplevel()
	okno_stol.focus_set()
	okno_stol.focus_force()
	okno_stol.geometry('500x360')
	okno_stol.resizable(False, False)
	okno_stol.title('Stolice')

	def cls_(nazwa_okna):
		for widgety in nazwa_okna.winfo_children():
			widgety.pack_forget()

	def cls_bez_zamknij(nazwa_okna, bez_zamknij):
		for widgety in nazwa_okna.winfo_children():
			if widgety is not bez_zamknij:
				widgety.pack_forget()



	kraj={

'Polski':'Warszawa',
'Niemiec':'Berlin',
'USA':'Waszyngton',
'Hiszpanii':'Madryt',
'Szwecji':'Sztokholm',
'Czech':'Praga',
'Ukrainy':'Kijów',
'Anglii':'Londyn',
'Meksyku':'Meksyk',
'Japonii':'Tokio',
'Korei Południowej':'Seul',
'Egiptu':'Kair',
'Francji':'Paryż',
'Holandii':'Amsterdam',
'Belgii':'Bruksela',
'Danii':'Kopenhaga',
'Włoch':'Rzym',
'Litwy':'Wilno',
'Norwegii':'Oslo',
'Austrii':'Wiedeń',
'Wenezueli':'Caracas',
'Wietnamu':'Hanoi',
'Turcji':'Ankara',
'RPA':'Pretoria',
'Argentyny':'Buenos Aires',
'Węgier':'Budapeszt',
'Rumunii':'Bukareszt',
'Kanady':'Ottawa',
'Portugalii':'Lizbona',
'Grecji':'Ateny',
'Białorusi':'Mińsk',
'Australii':'Canberra',
'Tajlandii':'Bangkok',
'Chin':'Pekin',
'Arabii Saudyjskiej':'Rijad',
'Iraku':'Bagdad',
'Iranu':'Teheran',
'Izraela':'Jerozolima',
'Rosji':'Moskwa',
'Syrii':'Damaszek',
'Jordanii':'Amman',
'Kolumbii':'Bogota',
'Boliwii':'La Paz',
'Szkocji':'Edynburg',
'Łotwy':'Ryga',
'Finlandii':'Helsinki',
'Słowacji':'Bratysława',
'Maroko':'Rabat',
'Peru':'Lima',
'Chorwacji': 'Zagrzeb',
'Słowenii': 'Lublana',
'Bułgarii': 'Sofia',
'Serbii': 'Belgrad',
'Bośni i Hercegowiny': 'Sarajewo',
'Estonii': 'Tallin',
'Mołdawii': 'Kiszyniów',
'Kosowa': 'Prisztina',
'Albanii': 'Tirana',
'Macedonii Północnej': 'Skopje',
'Czarnogóry': 'Podgorica',
'Malty': 'Valletta',
'Islandii': 'Rejkiawik',
'Liechtensteinu': 'Vaduz',
'Gruzji': 'Tbilisi',
'Mongolii': 'Ułan Bator',
'Nigerii': 'Abudża',
'Pakistanu': 'Islamabad'

}


#	 GRA GŁÓWNA

	def start(x):
		global count

		count+=1
		if count<21:
			x.destroy()
			ramka_gra=Frame(okno_stol)
			ramka_gra.pack()
			panstwo,stol=choice(list(kraj.items()))

			def check():
				odp=gra_odpowiedz.get()
				global pkt
				dobrze=choice(['Prawidłowo!','Zgadza się!','Tak!','Brawo!','Dobrze!',\
					'Świetnie!', 'Doskonale!', 'Rewelacyjnie!', 'Fantastycznie!', 'Super!',\
					'Bezapelacyjnie!', 'Oczywiście!', 'Jasne!', 'Absolutnie tak!',\
					'Bez wątpienia!', 'Okej!', 'Nieźle!'])
				zle=choice(['Źle!','Odpowiedź nieprawidłowa!','Nie!',\
					'Niestety nie.', 'Błąd :/', 'Pudło...', 'Niestety, to nie to :(', 'Nie tym razem.'])
				del kraj[panstwo]
				if odp.lower()==stol.lower():
					pkt+=1
					gra_pytanie.config(text=f'{dobrze}\n{count} z 20\nDotychczasowy wynik to {pkt} pkt.',\
						fg='green', font=('Arial', 16))
				else:
					pkt+=0
					gra_pytanie.config(text=f'{zle}\n{count} z 20\nDotychczasowy wynik to {pkt} pkt.', \
						fg='red', font=('Arial', 16))
					gra_popr=Label(ramka_gra,text=f'Stolica {panstwo} to {stol}!',font=('Arial', 14))
					gra_popr.pack(pady=5)
				gra_odpowiedz.config(state='disabled')
				potwierdz.config(text='Następna stolica', command=lambda:start(ramka_gra))
				



			gra_pytanie=Label(ramka_gra,text=f'Stolica {panstwo} to:', font=('Arial', 18))
			gra_pytanie.pack(pady=20)

			gra_odpowiedz=Entry(ramka_gra, justify='center', font=('Arial', 18))
			gra_odpowiedz.pack(pady=15)
			gra_odpowiedz.focus_set()

			potwierdz=Button(ramka_gra,text='Potwierdź', command=check, font=('Arial', 18))
			potwierdz.pack(side='bottom')



#	ZAKOŃCZENIE:

		else:
			x.destroy()
			ramka_gra=Frame(okno_stol)
			ramka_gra.pack()
			zakonczenie=Label(ramka_gra,text=f'Koniec!\
\nTwój wynik to: {pkt} pkt.', font=('Arial', 15)).pack(pady=15)

			if wyjscie_danych3()==None or wyjscie_danych2()==None or int(wyjscie_danych2())<int(pkt):
				zakonczenie2=Label(ramka_gra,font=('Arial',15),text=f'Rewelacja {wyjscie_danych1()}!\
\nTo najlepszy dotychczasowy wynik!').pack(pady=5)
				wprowadzenie_danych2(str(pkt))
				wprowadzenie_danych3(str(wyjscie_danych1()))
				ustal_rank(pkt)

			elif int(wyjscie_danych2())==int(pkt):
				zakonczenie2=Label(ramka_gra,font=('Arial',15),text=f'Brawo {wyjscie_danych1()}\
\n{pkt}pkt. to najwyższy wynik!\nZdobywasz tyle samo punktów co {wyjscie_danych3()}!!').pack(pady=5)
				ustal_rank(pkt)

			elif int(wyjscie_danych2())>int(pkt):
				zakonczenie2=Label(ramka_gra,font=('Arial',15),text=f'Najlepszy wynik dotychczas to:\
\n{wyjscie_danych2()}pkt.\nOsiągnięte przez użytkownika: {wyjscie_danych3()}').pack(pady=15)
				ustal_rank(pkt)



# WYŚWIETL NAJLEPSZE WYNIKI:

			def best():
				baza_danych=load_workbook('baza.xlsx')
				arkusz=baza_danych.active
				wyniki.config(state='disabled')
				g1=arkusz['B11'].value
				g2=arkusz['B12'].value
				g3=arkusz['B13'].value
				g4=arkusz['B14'].value
				g5=arkusz['B15'].value

				p1=arkusz['A11'].value
				p2=arkusz['A12'].value
				p3=arkusz['A13'].value
				p4=arkusz['A14'].value
				p5=arkusz['A15'].value


				okno_info=Tk()
				okno_info.geometry('300x200')
				okno_info.title('Highscores')
				baza_danych=load_workbook('baza.xlsx')
				info_napis=Label(okno_info,padx=20, text=f'Najlepsze dotychczasowe wyniki:\
\n\
\n1. {g1}   {p1}pkt.\
\n2. {g2}   {p2}pkt.\
\n3. {g3}   {p3}pkt.\
\n4. {g4}   {p4}pkt.\
\n5. {g5}   {p5}pkt.', font=('Arial', 12))
				info_napis.pack()
				baza_danych.close()


				def powrot_infoF():
					wyniki.config(state='active')
					okno_info.quit()
					okno_info.destroy()
				powrot_info=Button(okno_info,text='OK',command=powrot_infoF, font=('Arial', 18))
				powrot_info.pack(pady=15, ipadx=25)
				okno_info.mainloop()


			wyniki=Button(ramka_gra,text='Wyświetl najlepsze wyniki', command=best, font=('Arial', 18))
			wyniki.pack(pady=10)



	global count
	count=0
	global pkt
	pkt=0

	ramka_powitalna=Frame(okno_stol)
	ramka_powitalna.pack()

	powitanie=Label(ramka_powitalna,text=f'Witaj {wyjscie_danych1()}!\n\
\nGra polega na odgadnięciu jak największej liczby stolic.\n\
Jeśli odgadniesz wszystkie, zdobywasz 20 punktów.\n\
Powodzenia!', font=('Arial', 15)).pack(pady=20)

	dalej=Button(ramka_powitalna, text='Zaczynamy!',font=('Arial', 15),\
		command=lambda:start(ramka_powitalna)).pack(pady=10)





# PĘTLA OKNA I WYJŚCIE Z OKNA

	def powrot():
		okno_stol.destroy()
		okno.deiconify()

	zamknij=Button(okno_stol,text='Powrót do menu', borderwidth=4, command=powrot, font=("Arial",18))
	zamknij.pack(padx=15,pady=15,side='bottom')
	okno_stol.mainloop()

